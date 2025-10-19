import pandas as pd
import streamlit as st
import io
from openpyxl.styles import PatternFill

excel_df = st.file_uploader('Carregue o XLSX', type='xlsx')
opcao = st.selectbox('Qual profissão você deseja marcar?',
                     ('Engenheiro', 'Arquiteto', 'Desenvolvedor','Piloto','Recrutador'))

if excel_df is not None:
    @st.cache_data
    def carregar_dataframe():
        df_teste = pd.read_excel(excel_df, sheet_name=None)

        return df_teste

    df_teste = carregar_dataframe()

    st.bar_chart(df_teste['Profissões'], x='Profissão', y='Salário')

    @st.cache_data
    def formatar_doc(df, profissao_marcada):
        df_juntado = pd.merge(df['Profissões'], df['Prioridade'], on='Profissão', how='inner')
        idx_profissao = df_juntado[df_juntado['Profissão'] == profissao_marcada].index
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        nome_da_aba = 'Formatado'
        df_juntado.to_excel(writer, index=False, sheet_name=nome_da_aba)
        wb = writer.book
        ws = writer.sheets[nome_da_aba]
        cor_fundo = PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid")

        for pd_index in idx_profissao:
            excel_linha_num = pd_index + 2

            for celula in ws[excel_linha_num]:
                celula.fill = cor_fundo

        writer.close()

        data_bytes = output.getvalue()
        return data_bytes

    st.download_button(
        label='Baixe o documento formatado',
        data=formatar_doc(df_teste, opcao),
        file_name='doc.xlsx',
        icon=":material/download:",
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )